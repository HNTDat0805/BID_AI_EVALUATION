from __future__ import annotations

from pathlib import Path


def extract_text(file_path: str) -> str:
    """
    Extract text from a local file.

    Supported:
    - Images (.png, .jpg, .jpeg) via pytesseract
    - PDFs (.pdf) via pdfplumber
    - DOCX (.docx) via python-docx
    - XLSX (.xlsx) via openpyxl
    """

    path = Path(file_path)
    suffix = path.suffix.lower().lstrip(".")

    if suffix in {"png", "jpg", "jpeg"}:
        from PIL import Image
        import pytesseract

        image = Image.open(path)
        return pytesseract.image_to_string(image) or ""

    if suffix == "pdf":
        import pdfplumber

        extracted_parts: list[str] = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                extracted_parts.append(text)
        return "\n".join(extracted_parts).strip()

    if suffix == "docx":
        from docx import Document

        doc = Document(str(path))
        parts = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
        return "\n".join(parts).strip()

    if suffix == "xlsx":
        from openpyxl import load_workbook

        # read_only reduces memory usage for large sheets.
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append("\t".join(cells))
        return "\n".join(lines).strip()

    raise ValueError(f"Unsupported file extension: .{suffix}")

